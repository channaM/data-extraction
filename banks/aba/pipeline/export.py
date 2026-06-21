"""
Export a list of standard records to a clean, formatted .xlsx file.

Output path:
    output/{year}/{YYYYMM}/{CURRENCY}/aba_{YYYYMM}_{CURRENCY}.xlsx

The Excel file contains:
  - A styled header row (colored, bold, frozen)
  - Auto-fitted column widths
  - Number formatting for amounts
  - Date formatting for dates
  - Alternating row colors for readability
  - A proper Excel Table for easy filtering
"""

from __future__ import annotations

import logging
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from config import BANK_FULL_NAME, BANK_NAME

logger = logging.getLogger(__name__)

# ── Colour palette ────────────────────────────────────────────────────────────
COL_HEADER_BG   = "0F3460"   # deep navy
COL_HEADER_FG   = "FFFFFF"   # white
COL_ROW_ODD     = "F7FBFF"   # near-white blue tint
COL_ROW_EVEN    = "DDE9F5"   # light steel blue
COL_DEBIT_FG    = "C0392B"   # red for debit amounts
COL_CREDIT_FG   = "1A7A4A"   # green for credit amounts
COL_TITLE_BG    = "1A1A2E"   # dark navy for title row
COL_TITLE_FG    = "F0A500"   # gold text for title

# ── Column definitions ────────────────────────────────────────────────────────
COLUMNS = [
    ("Transaction Date", "transaction_date",  14, "date"),
    ("Value Date",       "value_date",        12, "date"),
    ("Reference No",     "reference_id",      18, "text"),
    ("Description",      "description",       38, "text"),
    ("Debit",            "debit_amount",      14, "amount_debit"),
    ("Credit",           "credit_amount",     14, "amount_credit"),
    ("Balance",          "balance",           16, "amount"),
]

DATE_FMT   = "DD/MM/YYYY"
AMOUNT_FMT = '#,##0.00'


def export(records: list[dict], currency: str, year: int, month: int,
           output_root: str = "output") -> Path:
    """
    Write records to a clean .xlsx file.

    Returns the path of the created file.
    """
    out_dir  = Path(output_root) / str(year) / f"{year}{month:02d}" / currency.upper()
    out_dir.mkdir(parents=True, exist_ok=True)

    filename = f"aba_{year}{month:02d}_{currency.upper()}.xlsx"
    out_path = out_dir / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    _write_title_row(ws, currency, year, month)
    _write_header_row(ws)
    last_data_row = _write_data_rows(ws, records)
    _apply_table(ws, last_data_row)
    _set_column_widths(ws)
    _freeze_panes(ws)
    _set_print_options(ws)

    wb.save(str(out_path))
    logger.info("[export] Saved %d records → '%s'", len(records), out_path)
    return out_path


# ──────────────────────────────────────────────────────────────────────────────
# Private helpers
# ──────────────────────────────────────────────────────────────────────────────

def _write_title_row(ws, currency: str, year: int, month: int) -> None:
    from calendar import month_name
    title = (
        f"{BANK_FULL_NAME}  —  Account Statement  —  "
        f"{month_name[month]} {year}  —  {currency}"
    )
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=len(COLUMNS))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font      = Font(name="Calibri", bold=True, size=12, color=COL_TITLE_FG)
    cell.fill      = PatternFill("solid", fgColor=COL_TITLE_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24


def _write_header_row(ws) -> None:
    thin = Side(style="thin", color="C0CFE0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (label, _, _, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font      = Font(name="Calibri", bold=True, size=10, color=COL_HEADER_FG)
        cell.fill      = PatternFill("solid", fgColor=COL_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border    = border

    ws.row_dimensions[2].height = 20


def _write_data_rows(ws, records: list[dict]) -> int:
    thin   = Side(style="thin", color="D0DCE8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_offset, record in enumerate(records):
        row_num  = row_offset + 3   # data starts at row 3
        is_odd   = (row_offset % 2 == 0)
        bg_color = COL_ROW_ODD if is_odd else COL_ROW_EVEN

        for col_idx, (_, field, _, fmt) in enumerate(COLUMNS, start=1):
            value = record.get(field)

            if isinstance(value, Decimal):
                value = float(value)

            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.fill   = PatternFill("solid", fgColor=bg_color)
            cell.border = border
            cell.font   = Font(name="Calibri", size=9)

            if fmt == "date":
                cell.number_format = DATE_FMT
                cell.alignment     = Alignment(horizontal="center")
            elif fmt == "amount_debit":
                cell.number_format = AMOUNT_FMT
                cell.alignment     = Alignment(horizontal="right")
                if value:
                    cell.font = Font(name="Calibri", size=9, color=COL_DEBIT_FG)
            elif fmt == "amount_credit":
                cell.number_format = AMOUNT_FMT
                cell.alignment     = Alignment(horizontal="right")
                if value:
                    cell.font = Font(name="Calibri", size=9, color=COL_CREDIT_FG)
            elif fmt == "amount":
                cell.number_format = AMOUNT_FMT
                cell.alignment     = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    return 2 + len(records)   # last row number with data


def _apply_table(ws, last_row: int) -> None:
    if last_row < 3:
        return
    first_col_letter = get_column_letter(1)
    last_col_letter  = get_column_letter(len(COLUMNS))
    ref = f"{first_col_letter}2:{last_col_letter}{last_row}"

    tbl = Table(displayName="ABA_Transactions", ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws.add_table(tbl)


def _set_column_widths(ws) -> None:
    for col_idx, (_, _, width, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _freeze_panes(ws) -> None:
    ws.freeze_panes = "A3"  # freeze title + header rows


def _set_print_options(ws) -> None:
    ws.print_title_rows = "1:2"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"
