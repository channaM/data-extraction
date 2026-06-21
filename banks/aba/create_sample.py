"""
Creates a realistic sample ABA Bank statement in XLSX format.

Run once to generate:   downloads/aba_sample_202501_USD.xlsx

The file mirrors the real ABA Bank Cambodia statement layout:
  - Rows 1–6: Bank header info (account, name, period, currency)
  - Row  7:   Blank
  - Row  8:   Column headers
  - Rows 9+:  Transaction data
"""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

OUTPUT_PATH = Path(__file__).parent / "downloads" / "aba_sample_202501_USD.xlsx"


TRANSACTIONS = [
    ("03/01/2025", "03/01/2025", "Opening Balance B/F",           "OB-2025001",   None,       None,      10_000.00),
    ("03/01/2025", "03/01/2025", "Salary Credit - Jan 2025",      "SAL-000123",   None,       3_500.00,  13_500.00),
    ("04/01/2025", "04/01/2025", "ATM Withdrawal - Aeon Mall PP", "ATM-078543",   200.00,     None,      13_300.00),
    ("05/01/2025", "05/01/2025", "Online Transfer to WING",       "TRF-220451",   150.00,     None,      13_150.00),
    ("07/01/2025", "07/01/2025", "Utility Bill - EDC Cambodia",   "UTIL-003321",  85.50,      None,      13_064.50),
    ("08/01/2025", "08/01/2025", "POS Purchase - Lucky Supermart","POS-994123",   62.30,      None,      13_002.20),
    ("10/01/2025", "10/01/2025", "Inward Remittance - TT",        "REM-INT-0012", None,       1_200.00,  14_202.20),
    ("12/01/2025", "12/01/2025", "Loan Repayment - Home Loan",    "LNR-888201",   450.00,     None,      13_752.20),
    ("14/01/2025", "14/01/2025", "ATM Withdrawal - ABA ATM 009", "ATM-078599",   300.00,     None,      13_452.20),
    ("15/01/2025", "15/01/2025", "Transfer from Savings Account", "TRF-SAV-001",  None,       500.00,    13_952.20),
    ("16/01/2025", "16/01/2025", "POS - Cafe Amazon Phnom Penh", "POS-994301",   8.50,       None,      13_943.70),
    ("18/01/2025", "18/01/2025", "Insurance Premium - Infinity",  "INS-2025-018", 120.00,     None,      13_823.70),
    ("20/01/2025", "20/01/2025", "Online Shopping - iShop",       "POS-994502",   45.99,      None,      13_777.71),
    ("22/01/2025", "22/01/2025", "POS - Total Petrol Station",    "POS-994611",   55.00,      None,      13_722.71),
    ("25/01/2025", "25/01/2025", "Dividend Credit - ACLEDA",      "DIV-2025-001", None,       350.00,    14_072.71),
    ("26/01/2025", "26/01/2025", "ATM Withdrawal - ABA ATM 002", "ATM-078701",   500.00,     None,      13_572.71),
    ("27/01/2025", "27/01/2025", "Interbank Transfer Out",        "IBK-OUT-0332", 1_000.00,   None,      12_572.71),
    ("28/01/2025", "28/01/2025", "POS - Makro Cambodia",          "POS-994788",   210.40,     None,      12_362.31),
    ("29/01/2025", "29/01/2025", "Loan Top-up Credit",            "LNR-TOP-0019", None,       2_000.00,  14_362.31),
    ("31/01/2025", "31/01/2025", "Bank Service Charge Jan 2025",  "SVC-2025-001", 3.00,       None,      14_359.31),
]

HEADER_ROWS = [
    ("ABA Bank",                     None),
    ("Account Statement",            None),
    ("Account No:",                  "000-123-456-789"),
    ("Account Name:",                "JOHN SMITH"),
    ("Currency:",                    "USD"),
    ("Period:",                      "01/01/2025 - 31/01/2025"),
]

COL_HEADERS = [
    "Date", "Value Date", "Narration", "Reference No", "Debit", "Credit", "Balance"
]

NAVY   = "1A1A2E"
GOLD   = "F0A500"
WHITE  = "FFFFFF"
LIGHT  = "EAF2FF"
BLUE   = "0F3460"


def _thin_border():
    s = Side(style="thin", color="C0CFE0")
    return Border(left=s, right=s, top=s, bottom=s)


def create_sample():
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ── Bank header rows ──────────────────────────────────────────────────
    for r_idx, (label, value) in enumerate(HEADER_ROWS, start=1):
        cell_a = ws.cell(row=r_idx, column=1, value=label)
        cell_a.font = Font(name="Calibri", bold=True, size=11, color=GOLD)
        cell_a.fill = PatternFill("solid", fgColor=NAVY)
        cell_a.alignment = Alignment(horizontal="left", vertical="center")

        if value:
            cell_b = ws.cell(row=r_idx, column=2, value=value)
            cell_b.font = Font(name="Calibri", size=11, color=WHITE)
            cell_b.fill = PatternFill("solid", fgColor=NAVY)

        # Extend the dark fill across all 7 columns
        for c in range(1, 8):
            ws.cell(row=r_idx, column=c).fill = PatternFill("solid", fgColor=NAVY)

        ws.row_dimensions[r_idx].height = 18

    # ── Blank row 7 ───────────────────────────────────────────────────────
    ws.row_dimensions[7].height = 6

    # ── Column header row (row 8) ─────────────────────────────────────────
    for c_idx, header in enumerate(COL_HEADERS, start=1):
        cell = ws.cell(row=8, column=c_idx, value=header)
        cell.font      = Font(name="Calibri", bold=True, size=10, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()
    ws.row_dimensions[8].height = 20

    # ── Transaction rows ──────────────────────────────────────────────────
    for r_offset, txn in enumerate(TRANSACTIONS):
        row_num  = r_offset + 9
        is_odd   = (r_offset % 2 == 0)
        bg_color = "F7FBFF" if is_odd else "DDE9F5"

        date_str, val_date, narration, ref, debit, credit, balance = txn

        values = [date_str, val_date, narration, ref,
                  debit if debit else None,
                  credit if credit else None,
                  balance]

        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=c_idx, value=val)
            cell.fill   = PatternFill("solid", fgColor=bg_color)
            cell.border = _thin_border()
            cell.font   = Font(name="Calibri", size=9)

            if c_idx in (1, 2):
                cell.alignment = Alignment(horizontal="center")
            elif c_idx in (5, 6, 7):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0.00"
                if c_idx == 5 and val:
                    cell.font = Font(name="Calibri", size=9, color="C0392B")
                elif c_idx == 6 and val:
                    cell.font = Font(name="Calibri", size=9, color="1A7A4A")
            else:
                cell.alignment = Alignment(horizontal="left")

        ws.row_dimensions[row_num].height = 15

    # ── Column widths ─────────────────────────────────────────────────────
    widths = [13, 13, 42, 20, 13, 13, 14]
    for c_idx, w in enumerate(widths, start=1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    # ── Freeze panes below the header row ─────────────────────────────────
    ws.freeze_panes = "A9"

    wb.save(str(OUTPUT_PATH))
    print(f"Sample created: {OUTPUT_PATH}")


if __name__ == "__main__":
    create_sample()
