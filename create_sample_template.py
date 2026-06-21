"""Run this once to generate a sample Excel template in templates/data_template.xlsx."""
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADERS = ["id", "name", "category", "price", "quantity", "description", "date"]

SAMPLE_ROWS = [
    ["P001", "Widget Alpha",   "Electronics", 29.99,  100, "High-quality widget",    "2026-01-15"],
    ["P002", "Gadget Beta",    "Electronics", 49.50,   50, "Latest gadget model",    "2026-02-20"],
    ["P003", "Donut Gamma",    "Food",         2.75, 1000, "Fresh daily donuts",     "2026-03-01"],
    ["P004", "Chair Delta",    "Furniture",  149.00,   30, "Ergonomic office chair", "2026-03-10"],
    ["P005", "Notebook Epsilon","Stationery",   3.99,  500, "A5 lined notebook",     "2026-04-05"],
    ["P006", "Lamp Zeta",      "Furniture",   39.99,   75, "LED desk lamp",          "2026-04-18"],
]

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
COL_WIDTHS   = [10, 22, 15, 10, 10, 30, 14]


def create_template(output_path: str = "templates/data_template.xlsx") -> None:
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Header row
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_idx - 1]

    ws.row_dimensions[1].height = 22

    # Sample data rows
    for row_data in SAMPLE_ROWS:
        ws.append(row_data)

    # Add a second sheet showing the column reference
    ws_ref = wb.create_sheet("Column Reference")
    ws_ref.append(["Column", "Type", "Required", "Notes"])
    reference = [
        ("id",          "text",    "No",  "Unique identifier; used for upsert"),
        ("name",        "text",    "Yes", "Display name of the record"),
        ("category",    "text",    "No",  "Grouping category"),
        ("price",       "decimal", "No",  "Numeric price (e.g. 29.99)"),
        ("quantity",    "integer", "No",  "Whole number quantity"),
        ("description", "text",    "No",  "Free-text description"),
        ("date",        "date",    "No",  "YYYY-MM-DD or DD/MM/YYYY"),
    ]
    for row in reference:
        ws_ref.append(row)

    wb.save(output_path)
    print(f"Template created: {output_path}")


if __name__ == "__main__":
    create_template()
